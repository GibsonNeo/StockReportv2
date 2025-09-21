#!/usr/bin/env python3
"""
build_score_report.py

Combines two CSVs on a key and computes scores for selected columns using one of two modes:
  1) quartile, bucketed grades with optional decile nudge
  2) standardized, continuous scores using zscore, robust_z, percentile, or minmax

Writes Excel or CSV. If Excel, can optionally apply conditional formatting by quartiles.

Also supports config-driven inputs and outputs. If CLI flags are omitted, the script reads grading.yml,
discovers input files, and writes the report to the configured reports directory.
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

import pandas as pd
import numpy as np
import yaml

# -------------------- util --------------------

def read_config(path: Path) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg

def find_config(cli_config: Optional[str]) -> Path:
    if cli_config:
        p = Path(cli_config)
        if not p.exists():
            sys.stderr.write(f"Config not found at {p}\n")
            sys.exit(2)
        return p
    cwd_yml = Path("grading.yml")
    if cwd_yml.exists():
        return cwd_yml
    script_dir_yml = Path(__file__).resolve().parent / "grading.yml"
    if script_dir_yml.exists():
        return script_dir_yml
    sys.stderr.write("No --config provided and grading.yml not found in CWD or script folder\n")
    sys.exit(2)

def _upper_strip(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.upper()

def _coerce_numeric(s: pd.Series) -> pd.Series:
    s2 = pd.to_numeric(s.astype(str).str.replace("%","", regex=False).str.strip(), errors="coerce")
    return s2

def resolve_input_path(io_cfg: Dict[str, Any]) -> Path:
    path = io_cfg.get("path")
    if path:
        p = Path(path)
        if not p.exists():
            sys.stderr.write(f"Input path does not exist: {p}\n")
            sys.exit(2)
        return p

    dir_ = io_cfg.get("dir", ".")
    pattern = io_cfg.get("pattern", "*.csv")
    pick = str(io_cfg.get("pick", "latest")).lower()

    base = Path(dir_)
    if not base.exists():
        sys.stderr.write(f"Input directory does not exist: {base}\n")
        sys.exit(2)

    matches = sorted(base.glob(pattern))
    if not matches:
        sys.stderr.write(f"No files matched pattern '{pattern}' in {base}\n")
        sys.exit(2)

    if pick == "first":
        return matches[0]
    return max(matches, key=lambda p: p.stat().st_mtime)

def resolve_output_path(io_cfg: Dict[str, Any]) -> Path:
    out_cfg = io_cfg.get("output", {})
    dir_ = out_cfg.get("dir", "./reports")
    name = out_cfg.get("filename", "Combined-Score-Report.xlsx")
    out_dir = Path(dir_)
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / name

# -------------------- quartile mode helpers --------------------

def _qcut_safe(values: pd.Series, q: int) -> Tuple[pd.Series, bool]:
    idx = values.index
    x = values.dropna()
    used_fallback = False
    if len(x) == 0:
        return pd.Series(np.nan, index=idx), used_fallback
    try:
        labels = pd.qcut(x, q=q, labels=list(range(1, q+1)), duplicates="drop")
        if len(pd.unique(labels)) < q:
            raise ValueError("qcut dropped bins")
    except Exception:
        used_fallback = True
        r = x.rank(method="average", ascending=True)
        edges = np.linspace(r.min(), r.max(), num=q+1)
        labels = pd.cut(r, bins=edges, include_lowest=True, labels=list(range(1, q+1)))
    out = pd.Series(np.nan, index=idx, dtype="float")
    out.loc[labels.index] = labels.astype(float)
    return out, used_fallback

def _decile_masks(values: pd.Series, higher_is_better: bool) -> Tuple[pd.Series, pd.Series]:
    x = values.dropna()
    if len(x) == 0:
        idx = values.index
        return pd.Series(False, index=idx), pd.Series(False, index=idx)
    lo_q = x.quantile(0.10)
    hi_q = x.quantile(0.90)
    if higher_is_better:
        top = values >= hi_q
        bottom = values <= lo_q
    else:
        top = values <= lo_q
        bottom = values >= hi_q
    return top.fillna(False), bottom.fillna(False)

def grade_quartile(values: pd.Series, higher_is_better: bool, n_quartiles: int, decile_nudge: float) -> pd.DataFrame:
    q_raw, used_fallback = _qcut_safe(values, n_quartiles)
    if higher_is_better:
        q_grade = q_raw.astype(float)
    else:
        q_grade = q_raw.map(lambda q: np.nan if pd.isna(q) else (n_quartiles + 1 - int(q))).astype(float)
    top_mask, bot_mask = _decile_masks(values, higher_is_better=higher_is_better)
    nudge = pd.Series(0.0, index=values.index, dtype="float")
    nudge[top_mask] = decile_nudge
    nudge[bot_mask] = -decile_nudge
    col_grade = (q_grade + nudge).clip(lower=1.0, upper=float(n_quartiles)).round(2)
    return pd.DataFrame({
        "quartile_raw": q_raw.astype("float"),
        "quartile_grade": q_grade.astype("float"),
        "decile_nudge": nudge,
        "column_grade": col_grade,
        "quartile_fallback_used": bool(used_fallback)
    }, index=values.index)

# -------------------- standardized mode helpers --------------------

def _winsorize_by_pct(values: pd.Series, p_low: float, p_high: float) -> pd.Series:
    x = values.copy()
    lo = x.quantile(p_low)
    hi = x.quantile(p_high)
    return x.clip(lower=lo, upper=hi)

def _std_z(values: pd.Series) -> pd.Series:
    mu = values.mean()
    sd = values.std(ddof=0)
    if pd.isna(sd) or sd == 0:
        return pd.Series(np.nan, index=values.index)
    return (values - mu) / sd

def _robust_z(values: pd.Series) -> pd.Series:
    med = values.median()
    mad = (values - med).abs().median()
    if pd.isna(mad) or mad == 0:
        return _std_z(values)
    return (values - med) / (mad * 1.4826)

def _percentile_rank(values: pd.Series) -> pd.Series:
    return values.rank(method="average", pct=True)

def _minmax_pct(values: pd.Series, p_low: float, p_high: float) -> pd.Series:
    lo = values.quantile(p_low)
    hi = values.quantile(p_high)
    if hi == lo:
        return pd.Series(0.5, index=values.index)
    return ((values - lo) / (hi - lo)).clip(0, 1)

def standardize_one(values: pd.Series, method: str, clip_z: float, pct_floor: float, pct_ceiling: float, transform: str) -> pd.Series:
    x = values.copy()
    if transform == "log1p":
        x = x.where(x >= 0, np.nan)
        x = np.log1p(x)
    elif transform == "abs":
        x = x.abs()
    elif transform == "negate":
        x = -x
    if method == "zscore":
        s = _std_z(x)
        if clip_z is not None:
            s = s.clip(-clip_z, clip_z)
        return s
    elif method == "robust_z":
        s = _robust_z(x)
        if clip_z is not None:
            s = s.clip(-clip_z, clip_z)
        return s
    elif method == "percentile":
        xw = _winsorize_by_pct(x, pct_floor, pct_ceiling)
        return _percentile_rank(xw)
    elif method == "minmax":
        xw = _winsorize_by_pct(x, pct_floor, pct_ceiling)
        return _minmax_pct(xw, pct_floor, pct_ceiling)
    else:
        raise ValueError(f"Unknown method {method}")

def combine_standardized(scores: Dict[str, pd.Series], weights: Dict[str, float], map_to: str):
    idx = None
    for s in scores.values():
        idx = s.index if idx is None else idx
    weighted_sum = pd.Series(0.0, index=idx, dtype="float")
    weight_present = pd.Series(0.0, index=idx, dtype="float")
    for name, s in scores.items():
        w = float(weights.get(name, 0.0))
        mask = s.notna()
        weighted_sum.loc[mask] += s[mask] * w
        weight_present.loc[mask] += w
    combined_raw = weighted_sum / weight_present.replace(0.0, np.nan)
    map_to = map_to or "percentile_0_100"
    if map_to == "linear_0_100":
        out = (combined_raw * 10.0) + 50.0
        out = out.clip(0.0, 100.0)
    elif map_to == "percentile_0_100":
        cr = combined_raw.dropna()
        if len(cr) == 0:
            out = pd.Series(np.nan, index=combined_raw.index)
        else:
            ranks = cr.rank(method="average", pct=True) * 100.0
            out = pd.Series(np.nan, index=combined_raw.index)
            out.loc[cr.index] = ranks
    else:
        raise ValueError(f"Unknown map_to {map_to}")
    return combined_raw, out, weight_present

# -------------------- main grading pipeline --------------------

def compute_quartile_scores(df: pd.DataFrame, spec: List[Dict[str, Any]], n_quartiles: int, decile_nudge: float) -> pd.DataFrame:
    weighted_sum = pd.Series(0.0, index=df.index, dtype="float")
    weight_present = pd.Series(0.0, index=df.index, dtype="float")
    parts = []
    for item in spec:
        name = item["name"]
        src = item["source_column"]
        direction = item.get("direction", "higher_better")
        weight = float(item.get("weight", 1.0))
        higher_is_better = direction == "higher_better"
        vals = _coerce_numeric(df[src]) if src in df.columns else pd.Series(np.nan, index=df.index, dtype="float")
        g = grade_quartile(vals, higher_is_better=higher_is_better, n_quartiles=n_quartiles, decile_nudge=decile_nudge)
        g = g.rename(columns={
            "quartile_raw": f"{name}_quartile_raw",
            "quartile_grade": f"{name}_quartile_grade",
            "decile_nudge": f"{name}_decile_nudge",
            "column_grade": f"{name}_grade_q",
            "quartile_fallback_used": f"{name}_quartile_fallback_used",
        })
        parts.append(g)
        mask = g[f"{name}_grade_q"].notna()
        weighted_sum.loc[mask] += g.loc[mask, f"{name}_grade_q"] * weight
        weight_present.loc[mask] += weight
    out = pd.concat([df] + parts, axis=1)
    combined_1_to_4 = weighted_sum / weight_present.replace(0.0, np.nan)
    combined_percent = ((combined_1_to_4 - 1.0) / 3.0) * 100.0
    out["combined_q_weighted_1_to_4"] = combined_1_to_4.round(3)
    out["combined_q_percent_0_to_100"] = combined_percent.round(2)
    out["combined_q_weights_covered"] = weight_present
    return out

def compute_standardized_scores(df: pd.DataFrame, spec: List[Dict[str, Any]], std_cfg: Dict[str, Any], map_to: str) -> pd.DataFrame:
    default_method = std_cfg.get("default_method", "zscore")
    default_clip_z = std_cfg.get("default_clip_z", 3.0)
    default_pct_floor = std_cfg.get("default_pct_floor", 0.01)
    default_pct_ceiling = std_cfg.get("default_pct_ceiling", 0.99)
    standardized = {}
    weights = {}
    parts = []
    for item in spec:
        name = item["name"]
        src = item["source_column"]
        direction = item.get("direction", "higher_better")
        weight = float(item.get("weight", 1.0))
        method = item.get("method", default_method)
        clip_z = float(item.get("clip_z", default_clip_z)) if item.get("clip_z", default_clip_z) is not None else None
        pct_floor = float(item.get("pct_floor", default_pct_floor))
        pct_ceiling = float(item.get("pct_ceiling", default_pct_ceiling))
        transform = item.get("transform", "none")
        vals = _coerce_numeric(df[src]) if src in df.columns else pd.Series(np.nan, index=df.index, dtype="float")
        s = standardize_one(vals, method=method, clip_z=clip_z, pct_floor=pct_floor, pct_ceiling=pct_ceiling, transform=transform)
        if direction == "lower_better":
            s = -s
        standardized[name] = s
        weights[name] = weight
        parts.append(pd.DataFrame({f"{name}_standardized": s}, index=df.index))
    combined_raw, combined_0_100, weights_cov = combine_standardized(standardized, weights, map_to=map_to)
    out = pd.concat([df] + parts, axis=1)
    out["combined_z_weighted"] = combined_raw
    out["combined_z_percent_0_to_100"] = combined_0_100
    out["combined_z_weights_covered"] = weights_cov
    return out

# -------------------- formatting helpers --------------------

def reorder_after_column(df: pd.DataFrame, target_col: str, insert_col: str) -> pd.DataFrame:
    if insert_col not in df.columns:
        return df
    if target_col in df.columns:
        cols = list(df.columns)
        cols.remove(insert_col)
        idx = cols.index(target_col) + 1
        cols.insert(idx, insert_col)
        return df[cols]
    lowered = {c.lower(): c for c in df.columns}
    if target_col.lower() in lowered:
        anchor = lowered[target_col.lower()]
        cols = list(df.columns)
        cols.remove(insert_col)
        idx = cols.index(anchor) + 1
        cols.insert(idx, insert_col)
        return df[cols]
    return df

def apply_excel_formatting(path: Path, df: pd.DataFrame, cfg: Dict[str, Any]):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    def _apply_transform(series: pd.Series, transform: str) -> pd.Series:
        s = pd.to_numeric(series, errors="coerce")
        if transform == "log1p":
            s = s.where(s >= 0, np.nan)
            return np.log1p(s)
        if transform == "abs":
            return s.abs()
        if transform == "negate":
            return -s
        return s

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    fmt_cfg = cfg.get("formatting", {})
    if not bool(fmt_cfg.get("enabled", True)):
        wb.save(path)
        return

    # Optional, hide all columns except a whitelist, case-insensitive match
    visible_cols_cfg = fmt_cfg.get("visible_columns", [])
    visible_cols = [c.strip() for c in visible_cols_cfg] if isinstance(visible_cols_cfg, list) else []
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_to_idx = {cell.value: cell.column for cell in header_row if cell.value is not None}

    if visible_cols:
        lowers = {str(h).lower(): h for h in header_to_idx.keys()}
        to_keep_idx = set()
        for want in visible_cols:
            key = str(want).lower()
            if key in lowers:
                to_keep_idx.add(header_to_idx[lowers[key]])
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            if col_idx not in to_keep_idx:
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Freeze header row
    if bool(fmt_cfg.get("freeze_header", True)):
        ws.freeze_panes = "A2"

    # Add AutoFilter across the full width
    if bool(fmt_cfg.get("add_autofilter", True)):
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Autosize columns based on header and sampled data
    if bool(fmt_cfg.get("autosize_columns", True)):
        sample_rows = int(fmt_cfg.get("autosize_sample_rows", 1000))
        pad = int(fmt_cfg.get("autosize_padding_chars", 2))
        widths = {}
        for cell in header_row:
            if cell.value is None:
                continue
            col = cell.column
            text = str(cell.value)
            widths[col] = max(widths.get(col, 0), len(text))
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 1 + sample_rows)):
            for cell in row:
                col = cell.column
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                if len(s) > 200:
                    s = s[:200]
                widths[col] = max(widths.get(col, 0), len(s))
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = max(6, w + pad)

    # Colors
    fill_green = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_blue  = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    fill_yell  = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    fill_red   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    graded = cfg.get("grade_columns", [])
    targets = []
    combined_col = "combined_z_percent_0_to_100"
    if combined_col in df.columns:
        targets.append((combined_col, "higher_better", None))
    for item in graded:
        src = item.get("source_column")
        direction = item.get("direction", "higher_better")
        transform = item.get("transform", "none")
        if src in df.columns:
            targets.append((src, direction, transform))

    def col_letter(col_name: str):
        idx = header_to_idx.get(col_name)
        if idx is None:
            return None
        return get_column_letter(idx)

    n_rows = df.shape[0]
    start_row = 2
    end_row = n_rows + 1

    for col_name, direction, transform in targets:
        letter = col_letter(col_name)
        if not letter:
            continue

        s = df[col_name]
        s_t = _apply_transform(s, transform or "none")
        q1 = s_t.quantile(0.25)
        q2 = s_t.quantile(0.50)
        q3 = s_t.quantile(0.75)

        rng = f"{letter}{start_row}:{letter}{end_row}"
        first_cell = f"{letter}{start_row}"

        if (transform or "none") == "abs":
            expr = f"ABS({first_cell})"
        elif (transform or "none") == "negate":
            expr = f"(-{first_cell})"
        elif (transform or "none") == "log1p":
            expr = f"LN(1+MAX({first_cell},0))"
        else:
            expr = first_cell

        top_fill, q3_fill, q2_fill, bottom_fill = fill_green, fill_blue, fill_yell, fill_red
        if direction == "lower_better":
            top_fill, q3_fill, q2_fill, bottom_fill = fill_red, fill_yell, fill_blue, fill_green

        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{expr}<={q1}'], stopIfTrue=False, fill=bottom_fill))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND({expr}>{q1},{expr}<={q2})'], stopIfTrue=False, fill=q2_fill))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND({expr}>{q2},{expr}<={q3})'], stopIfTrue=False, fill=q3_fill))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{expr}>{q3}'], stopIfTrue=False, fill=top_fill))

    wb.save(path)

# -------------------- main --------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-a", help="Path to first CSV")
    ap.add_argument("--input-b", help="Path to second CSV")
    ap.add_argument("--config", help="Path to grading .yml")
    ap.add_argument("--out", help="Output file, .xlsx or .csv")
    args = ap.parse_args()

    cfg_path = find_config(args.config)
    cfg = read_config(cfg_path)

    io_cfg = cfg.get("io", {})
    in_a = Path(args.input_a) if args.input_a else resolve_input_path(io_cfg.get("input_a", {}))
    in_b = Path(args.input_b) if args.input_b else resolve_input_path(io_cfg.get("input_b", {}))
    out_path = Path(args.out) if args.out else resolve_output_path(io_cfg)

    key = cfg.get("join_key", "Ticker")
    join_type = cfg.get("join_type", "inner")
    spec = cfg.get("grade_columns", [])
    mode = cfg.get("scoring", {}).get("mode", "standardized")

    df_a = pd.read_csv(in_a)
    df_b = pd.read_csv(in_b)

    if key not in df_a.columns or key not in df_b.columns:
        sys.stderr.write(f"Join key '{key}' must exist in both CSVs\n")
        sys.exit(2)

    df_a[key] = _upper_strip(df_a[key])
    df_b[key] = _upper_strip(df_b[key])

    if join_type == "inner":
        merged = pd.merge(df_a, df_b, on=key, how="inner", suffixes=("_A", "_B"))
    elif join_type == "left_a":
        merged = pd.merge(df_a, df_b, on=key, how="left", suffixes=("_A", "_B"))
    elif join_type == "left_b":
        merged = pd.merge(df_b, df_a, on=key, how="left", suffixes=("_B", "_A"))
    elif join_type == "outer":
        merged = pd.merge(df_a, df_b, on=key, how="outer", suffixes=("_A", "_B"))
    else:
        sys.stderr.write(f"Unsupported join_type '{join_type}'\n")
        sys.exit(2)

    out = merged.copy()

    if mode in ("quartile", "both"):
        n_quartiles = int(cfg.get("n_quartiles", 4))
        decile_nudge = float(cfg.get("decile_nudge", 0.25))
        out = compute_quartile_scores(out, spec=spec, n_quartiles=n_quartiles, decile_nudge=decile_nudge)

    if mode in ("standardized", "both"):
        std_cfg = cfg.get("standardization", {})
        map_to = cfg.get("scoring", {}).get("map_to", "percentile_0_100")
        out = compute_standardized_scores(out, spec=spec, std_cfg=std_cfg, map_to=map_to)

    insert_after = cfg.get("formatting", {}).get("insert_combined_after_column", "Sector")
    if "combined_z_percent_0_to_100" in out.columns and insert_after in out.columns:
        out = reorder_after_column(out, target_col=insert_after, insert_col="combined_z_percent_0_to_100")

    if out_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        try:
            import openpyxl  # noqa
        except Exception:
            sys.stderr.write("openpyxl is required to write Excel files\n")
            sys.exit(2)
        out.to_excel(out_path, index=False)
        apply_excel_formatting(out_path, out, cfg)
    else:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out.to_csv(out_path, index=False)

    print(f"Wrote {out_path} with {len(out)} rows")

if __name__ == "__main__":
    main()
